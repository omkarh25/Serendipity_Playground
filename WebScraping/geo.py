<<<<<<< HEAD
from geopy.geocoders import Nominatim
import pandas as pd

# Correctly formatted file path
file_path = r'C:\Users\91861\OneDrive\Desktop\bhoodevi\WebScraping\bhoodevi 2.xlsx'  # Raw string to handle file path

# Load the Excel file
addresses_df = pd.read_excel(file_path)

# Initialize geolocator
geolocator = Nominatim(user_agent="geoapiExercises")

# Function to geocode address
def geocode_address(address):
    try:
        location = geolocator.geocode(address)
        if location:
            return location.latitude, location.longitude
        else:
            return None, None
    except Exception as e:
        return None, None

# Geocode each address in the dataframe
addresses_df['Coordinates'] = addresses_df['Location'].apply(geocode_address)

# Splitting the coordinates into latitude and longitude
addresses_df[['Geo Latitude', 'Geo Longitude']] = pd.DataFrame(addresses_df['Coordinates'].tolist(), index=addresses_df.index)

# Save the updated results to a new Excel file
addresses_df.to_excel('geocoded_addresses_updated.xlsx', index=False)

print("Geocoding complete. Updated results saved to 'geocoded_addresses_updated.xlsx'.")
=======
import time
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderServiceError
from openpyxl import load_workbook

def do_geocode(address, geolocator):
    try:
        return geolocator.geocode(address)
    except GeocoderTimedOut:
        time.sleep(1)  # Retry after a short wait in case of timeout
        return do_geocode(address, geolocator)
    except GeocoderServiceError as e:
        print(f"Geocoder service error for address '{address}': {e}")
        return None

def main():
    file_path = r'C:\Users\91861\OneDrive\Desktop\bhoodevi\WebScraping\your_excel_file.xlsx'  # Update with your file name
    wb = load_workbook(file_path)
    sheet = wb.active
    geolocator = Nominatim(user_agent="geoapiExercises")

    for index, row in enumerate(sheet.iter_rows(min_row=1, max_col=1, values_only=True), start=1):
        if row[0] and isinstance(row[0], str) and row[0].strip():
            address = row[0]  # Assuming address is in the first column
            location = do_geocode(address, geolocator)
            if location:
                sheet.cell(row=index, column=2).value = location.latitude
                sheet.cell(row=index, column=3).value = location.longitude
            else:
                print(f"Location not found or error occurred for address: {address}")
        else:
            print(f"Empty or invalid address at row {index}")

    updated_file_path = r'C:\Users\91861\OneDrive\Desktop\bhoodevi\WebScraping\updated_your_excel_file.xlsx'
    wb.save(updated_file_path)  # Saves the workbook with coordinates

if __name__ == "__main__":
    main()
>>>>>>> 3a05ebf1f9e6b31c23aae6394e633958849758de
